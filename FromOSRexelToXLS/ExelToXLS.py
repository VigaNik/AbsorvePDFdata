import openpyxl
import re
import xml.etree.ElementTree as ET
import logging
from typing import List, Tuple, Optional

# Configuration
xlsx_path = r'C:/NikWorckSpase/Testingfiles/pdfToXML/26580476.xlsx'
output_xml_path = xlsx_path.replace(".xlsx", "_references.xml")
END_MARKER_TEXT = "This content downloaded from"
PAGE_PATTERN = r"\b(page\.?)\s*\d+"
TIME_PATTERN = r"\b\d{1,2}:\d{2}(:\d{2})?\b"

# Logging configuration
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

def contains_lowercase(text: str) -> bool:
    """Check if the text contains lowercase letters."""
    return any(char.islower() for char in text)

def is_smaller_than_main_text(current_font_size: Optional[int], main_text_font_size: Optional[int]) -> bool:
    """Check if current font size is smaller than the main text font size."""
    return current_font_size is not None and main_text_font_size is not None and current_font_size < main_text_font_size

def matches_start_pattern(text: str) -> bool:
    """Check if text matches the start pattern of a reference."""
    start_pattern = r"^\d+[\.\)]?\s+.*"
    return bool(re.match(start_pattern, text))

def contains_end_marker(text: str) -> bool:
    """Check if text contains the end marker."""
    return END_MARKER_TEXT in text

def matches_time_pattern(text: str) -> bool:
    """Check if text matches a time pattern (HH:MM or HH:MM:SS)."""
    return bool(re.search(TIME_PATTERN, text))

def extract_references_from_xlsx(xlsx_path: str) -> List[Tuple[int, str, str]]:
    """Extracts references from an XLSX file."""
    try:
        workbook = openpyxl.load_workbook(xlsx_path, data_only=False)
        logging.info(f"Opened workbook: {xlsx_path}")
    except Exception as e:
        logging.error(f"Error opening XLSX file: {e}")
        return []

    references = []
    current_ref_num = 1
    threshold_found = False
    main_text_font_size = None

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_count = sheet.max_row
        start_row = row_count // 2

        current_reference = []
        rows = list(sheet.iter_rows(min_row=start_row, values_only=False))
        logging.info(f"Processing sheet: {sheet_name}, starting from row {start_row}")

        for i, row in enumerate(rows):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    current_font_size = cell.font.size if cell.font and contains_lowercase(cell.value) else None
                    if main_text_font_size is None and current_font_size:
                        main_text_font_size = current_font_size


                    if (matches_start_pattern(cell.value) or
                        (is_smaller_than_main_text(current_font_size, main_text_font_size) and current_font_size != main_text_font_size) or
                        threshold_found):
                        threshold_found = True
                        current_reference.append(cell.value.strip())
                        logging.debug(f"Collecting reference text: '{cell.value.strip()}'")


                    if contains_end_marker(cell.value) or matches_time_pattern(cell.value) or cell.value.strip() == "":

                        if current_reference:
                            save_reference(current_reference, current_ref_num, references, sheet_name)
                            current_reference = []
                            current_ref_num += 1
                        threshold_found = False
                        if contains_end_marker(cell.value):
                            break

    workbook.close()
    logging.info(f"Extracted {len(references)} references from {xlsx_path}")
    return references

def save_reference(current_reference: List[str], current_ref_num: int, references: List[Tuple[int, str, str]], sheet_name: str):
    """Saves the current reference and appends it to the references list."""
    full_reference = " ".join(current_reference).strip()
    # Удаляем маркер end_marker из текста, если он присутствует
    if END_MARKER_TEXT in full_reference:
        full_reference = full_reference.split(END_MARKER_TEXT)[0].strip()
    references.append((current_ref_num, full_reference, sheet_name))
    logging.info(f"Saved reference #{current_ref_num} from sheet {sheet_name}")

def save_references_to_xml(references: List[Tuple[int, str, str]], output_xml_path: str):
    """Saves references to an XML file."""
    if not references:
        logging.warning("No references found to save.")
        return

    root = ET.Element("References")
    for ref_num, reference, page in references:
        ref_element = ET.SubElement(root, "Reference")
        ref_element.set("number", str(ref_num))
        ref_element.set("page", page)
        ref_element.text = reference

    try:
        tree = ET.ElementTree(root)
        tree.write(output_xml_path, encoding="utf-8", xml_declaration=True)
        logging.info(f"References successfully saved to {output_xml_path}")
    except Exception as e:
        logging.error(f"Error saving XML file: {e}")

if __name__ == "__main__":
    references = extract_references_from_xlsx(xlsx_path)
    if references:
        save_references_to_xml(references, output_xml_path)