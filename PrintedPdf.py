import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from bidi.algorithm import get_display
import re

pdf_text = r'C:/NikWorckSpase/Testingfiles/26580476 - Copy (6).pdf'
output_xlsx_path = pdf_text.replace(".pdf", "_styled_text.xlsx")

def extract_text_with_styles(pdf_text):
    """
    Extracts text, font size, font type, bold, and italic from each page in the given PDF file.
    Returns a list of dictionaries for each page, containing text and styling information.
    """
    doc = fitz.open(pdf_text)
    pages_data = []

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        page_data = []

        # Process each text block with font info
        for block in page.get_text("dict")["blocks"]:
            if "lines" in block:  # This block contains text
                block_text = ""  # Accumulate text with line breaks
                for line in block["lines"]:
                    for span in line["spans"]:
                        # Collect text, font size, font name, bold, and italic information
                        text = clean_text(span["text"])
                        font_size = span["size"]
                        font_name = span["font"]

                        # Determine if text is bold or italic based on the font name
                        is_bold = "Bold" in font_name or "bold" in font_name
                        is_italic = "Italic" in font_name or "italic" in font_name

                        # Adjust display for mixed Hebrew and English directions
                        #text = reorder_mixed_text(text)

                        # Append text with line break after each line
                        block_text += text + "\n"

                # Store each line with detailed styling info
                page_data.append({
                    "Text": block_text.strip(),  # Remove trailing newline
                    "FontSize": font_size,
                    "FontName": font_name,
                    "Bold": is_bold,
                    "Italic": is_italic
                })

        pages_data.append(page_data)

    return pages_data

def reorder_mixed_text(text):
    """
    Reorders mixed Hebrew and English text: Hebrew text is displayed RTL, English text LTR.
    """
    # Split text into words to separate Hebrew and English parts
    words = text.split()
    reordered_words = []

    for word in words:
        # Check if word contains Hebrew characters
        if re.search(r'[\u0590-\u05FF]', word):
            reordered_words.append(get_display(word[::-1]))  # Reverse Hebrew word
        else:
            reordered_words.append(word)  # English or other language, keep as is

    # Join words in reversed order to simulate reading from right to left
    return " ".join(reordered_words[::-1])

def clean_text(text):
    """
    Removes characters that are not allowed in Excel cells.
    """
    # Remove control characters and other non-printable characters
    return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)

def save_text_to_styled_xlsx(pages_data, output_xlsx_path):
    """
    Save text and styles to an Excel file with each page in a separate sheet.
    Sets entire sheet orientation to right-to-left for Hebrew text.
    """
    wb = Workbook()
    for i, page_data in enumerate(pages_data, start=1):
        ws = wb.create_sheet(title=f"Page_{i}")

        # Set the sheet orientation to right-to-left if needed
        ws.sheet_view.rightToLeft = True

        row_num = 1

        for line_data in page_data:
            # Write the text with font name, size, bold, and italic properties
            cell = ws.cell(row=row_num, column=1, value=line_data["Text"])
            cell.font = Font(
                name=line_data["FontName"],
                size=line_data["FontSize"],
                bold=line_data["Bold"],
                italic=line_data["Italic"]
            )
            cell.alignment = Alignment(horizontal="right", wrap_text=True)  # Right-align and wrap text

            row_num += 1  # Move to the next row

    # Remove default sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(output_xlsx_path)
    print(f"Text with styles (bold and italic) successfully saved to {output_xlsx_path}")

def extract_pdf_to_styled_excel():
    pages_data = extract_text_with_styles(pdf_text)
    save_text_to_styled_xlsx(pages_data, output_xlsx_path)

if __name__ == "__main__":
    extract_pdf_to_styled_excel()
